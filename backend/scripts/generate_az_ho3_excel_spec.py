import sys
from pathlib import Path

import openpyxl

# Ensure backend root is in sys.path when script is executed directly
backend_dir = Path(__file__).resolve().parent.parent
if str(backend_dir) not in sys.path:
    sys.path.insert(0, str(backend_dir))

from app.ipir.package import IPIRPackage


def generate_excel_spec(json_spec_path: Path, output_excel_path: Path) -> None:
    """Generates a deterministic synthetic Excel actuarial workbook from the actuarial rate spec."""
    with open(json_spec_path, encoding="utf-8") as f:
        pkg = IPIRPackage.model_validate_json(f.read())

    spec = pkg.model_dump(mode="json")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. Metadata Sheet
    ws_meta = wb.create_sheet(title="Metadata")
    ws_meta.append(["Property", "Value"])
    ws_meta.append(["id", pkg.id])
    ws_meta.append(["name", pkg.name])
    ws_meta.append(["version", pkg.version])
    ws_meta.append(["start_date", str(pkg.effective_period.start)])
    ws_meta.append(["product_id", pkg.product.id])
    ws_meta.append(["state", pkg.product.jurisdiction.state_or_province or "AZ"])

    # 2. Inputs Sheet
    ws_inputs = wb.create_sheet(title="Inputs")
    ws_inputs.append(["id", "name", "data_type", "minimum", "maximum", "allowed_values"])
    for inp in spec.get("inputs", []):
        ws_inputs.append([
            inp["id"],
            inp["name"],
            inp["data_type"],
            inp.get("minimum"),
            inp.get("maximum"),
            ",".join(inp["allowed_values"]) if inp.get("allowed_values") else "",
        ])

    # 3. Rate Table Sheets
    for table in spec.get("tables", []):
        t_id = table["id"]
        sheet_title = t_id[:31]  # Excel sheet title max length 31
        ws_t = wb.create_sheet(title=sheet_title)

        if len(table["dimensions"]) == 1:
            dim_ref = table["dimensions"][0]["input_ref"]
            ws_t.append([dim_ref, "Value"])
            for row in table["rows"]:
                m = row["matches"][0]
                if "value" in m and m["value"] is not None:
                    match_str = str(m["value"])
                else:
                    match_str = f"{m.get('minimum', '')}..{m.get('maximum', '')}"
                ws_t.append([match_str, row["value"]])
        elif len(table["dimensions"]) == 2:
            # 2D Table format
            ws_t.append(["dim1_ref", "dim1_val", "dim2_ref", "dim2_val", "value"])
            d1_ref = table["dimensions"][0]["input_ref"]
            d2_ref = table["dimensions"][1]["input_ref"]
            for row in table["rows"]:
                m1 = row["matches"][0]
                m2 = row["matches"][1]
                ws_t.append([
                    d1_ref,
                    str(m1.get("value", "")),
                    d2_ref,
                    str(m2.get("value", "")),
                    row["value"],
                ])

    # 4. Modifiers Sheet
    ws_mod = wb.create_sheet(title="Modifiers")
    ws_mod.append(["id", "name", "modifier_type", "applies_to", "value"])
    for mod in spec.get("modifiers", []):
        ws_mod.append([
            mod["id"],
            mod["name"],
            mod["modifier_type"],
            mod["applies_to"],
            mod["value"],
        ])

    # 5. Constraints Sheet
    ws_con = wb.create_sheet(title="Constraints")
    ws_con.append(["id", "name", "constraint_type", "amount", "applies_to", "sequence"])
    for con in spec.get("constraints", []):
        ws_con.append([
            con["id"],
            con["name"],
            con["constraint_type"],
            con["amount"],
            con["applies_to"],
            con.get("sequence"),
        ])

    # 6. Fees Sheet
    ws_fee = wb.create_sheet(title="Fees")
    ws_fee.append(["id", "name", "amount", "applies_to", "sequence"])
    for fee in spec.get("fees", []):
        ws_fee.append([
            fee["id"],
            fee["name"],
            fee["amount"],
            fee["applies_to"],
            fee.get("sequence"),
        ])

    # 7. Calculations Sheet
    ws_calc = wb.create_sheet(title="Calculations")
    ws_calc.append(["id", "name", "operator"])
    for calc in spec.get("calculations", []):
        ws_calc.append([
            calc["id"],
            calc["name"],
            calc["expression"].get("operator", "BINARY"),
        ])

    output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel_path)
    print(f"Successfully generated synthetic Excel rate spec at: {output_excel_path}")


def main() -> None:
    root_dir = Path(__file__).resolve().parent.parent.parent
    json_path = root_dir / "data" / "actuarial" / "AZ_HO3_2026_09_rate_spec.json"
    excel_path = root_dir / "data" / "actuarial" / "AZ_HO3_2026_09_rate_spec.xlsx"

    generate_excel_spec(json_path, excel_path)


if __name__ == "__main__":
    main()

